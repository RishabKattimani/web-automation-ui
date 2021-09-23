# ------------------------------------------------------------------------------
# Imports

import PySimpleGUI as sg
import requests
import time
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import random
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl



# ------------------------------------------------------------------------------
# Window Layout

sg.theme('DefaultNoMoreNagging')

font = ("Inter, 20")

layout = [
    [sg.Text('Excel Web Automation Time Application', size=(75, 1), font=("Helvetica", 45), text_color='black')],

    [sg.Text("                      ", size=(25,1), font=('Arial', 20), text_color='white')],

    [sg.Text('Set The Time For Submitting the Form Here!', size=(50,1), font=("Arial", 20), text_color='black')],

    [sg.Text("                     ", size=(25,1), font=('Arial', 20), text_color='white')],

    [sg.Text('Form URL: ', size=(10,1), font=("Arial", 20), text_color='black'), sg.InputText(size=(50,1), font=('Arial', 20))],


    [sg.Text("                     ", size=(25,1), font=('Arial', 20), text_color='white')],


    [sg.Radio('Submit form every : ', "RADIO1", default=True, size=(20,1),font=('Arial', 20)), sg.InputText(size=(10,1), font=('Arial', 20)), sg.Text("Minutes", font=('Arial', 20))],

    [sg.Radio('Submit form Randomly Between every : ', "RADIO1", font=('Arial', 20)),  sg.InputText(size=(10,1), font=('Arial', 20)), sg.Text(' minutes to  : ', font=('Arial', 20) ), sg.InputText(size=(10,1), font=('Arial', 20)), sg.Text("Minutes.", font=('Arial', 20))],

    [sg.Text("                     ", size=(25,1), font=('Arial', 20), text_color='white')],

    [sg.Text('Excel Data Filename', size=(18,1), font=("Arial", 20), text_color='black'), sg.InputText(size=(10,1), font=('Arial', 20))],

    [sg.Text("                     ", size=(25,1), font=('Arial', 20), text_color='white')],

    [sg.Text('Exit After ', size=(18,1), font=("Arial", 20), text_color='black'), sg.InputText(size=(10,1), font=('Arial', 20)), sg.Text('Times ', size=(18,1), font=("Arial", 20), text_color='black')],

    [sg.Text("                     ", size=(25,1), font=('Arial', 20), text_color='white')],

    [sg.Submit(size=(19,1), font=('Arial', 20)), sg.Cancel(size=(19,1), font=('Arial', 20))],

]

# ------------------------------------------------------------------------------
# Window

window = sg.Window('Automated Form Filling', layout, size=(1400, 600))
event, values = window.read()
window.close()

if event == 'Cancel' or event == None:
    quit()
#
# print('Value 0', values[0])
# print('Value 1', values[1])
# print('Value 2', values[2])
# print('Value 3', values[3])
# print('Value 4', values[4])
# print('Value 5', values[5])
# print('Value6', values[6])
print(values[7])
repeat_times = int(values[7])


# ------------------------------------------------------------------------------
# Getting ID / Xpath

url = values[0]
page = requests.get(url)
soup = BeautifulSoup(page.text, 'html.parser')

form_tags = soup.findAll("input", {"class": "lp-form-react__input js-input-mapping"})
button_tag = soup.find("button")

name_xpath = '//*[@id="' + form_tags[1]['id'] + '"]'

phone_xpath = '//*[@id="' + form_tags[2]['id'] + '"]'

address_xpath = '//*[@id="' + form_tags[3]['id'] + '"]'


# ------------------------------------------------------------------------------
# Excel Data

print("\n")
path = str(values[6])


try:

    wb_obj = openpyxl.load_workbook(path.strip())
    # from the active attribute
    sheet_obj = wb_obj.active

    # get max column count
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row

# ------------------------------------------------------------------------------
# Looping Through the Data

    for j in range(2, max_row+1):

        if (sheet_obj.cell(row=j,column=8).value == 'si'):

                    print("Data Already Submitted, So Skipping ", sheet_obj.cell(row=j,column=7).value)
                    time.sleep(2)
                    continue

        else:

            print("Submitting ", sheet_obj.cell(row=j,column=7).value)

# ------------------------------------------------------------------------------
# Submitting The Data

            driver = webdriver.Chrome(ChromeDriverManager().install()) # Open New Instance

            driver.get(values[0])

            time.sleep(2)

            name = driver.find_element_by_xpath(name_xpath)
            name.send_keys(sheet_obj.cell(row=j,column=7).value)

            phone = driver.find_element_by_xpath(phone_xpath)
            phone.send_keys(sheet_obj.cell(row=j,column=6).value)

            address = driver.find_element_by_xpath(address_xpath)
            address.send_keys(sheet_obj.cell(row=j,column=4).value)

            button = driver.find_element_by_tag_name('button').click()
            repeat_times = (repeat_times - 1)
            time.sleep(5)


            driver.quit()


            print(j)

# ------------------------------------------------------------------------------
# Saving The Data


            submitted = sheet_obj.cell(row=j,column=8)
            submitted.value =  'si'

            wb_obj.save(values[6])
            print('Submitted The Data')

# ------------------------------------------------------------------------------
# Timing



            if values[1] == True:  # If custom timing chosen:

                print('Chosen Duration: ', int(values[2]), 'Minutes')
                time.sleep(int(values[2]) * 60) # multiply by 60 later

            else: # If Range Timing Chosen

                random_generate = random.randint(int(values[4]), int(values[5])) # Randomly Generated Time
                print('Randomly Generated Duration: ', random_generate, 'Minutes')
                time.sleep(random_generate * 60)

            #
            if repeat_times <= 0:
                print('stopping')
                break


# ------------------------------------------------------------------------------
# Submitting The Data

except Exception as e:
    print(e)
    print ( "Error : Contact Us")
print("Success!")
